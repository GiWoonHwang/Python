# 사이트에서 이메일 정보를 수집하여 엑셀에 기록, 이메일의 정규식 표현으로 이메일 형식을 찾아 수집

# import re

# test_string = '''
# tjgustlr@tjgustlr.com
# tjgustlr@naver.com
# tjgustlr1@yahoo.com
# tjgustlr2@daum.com
# tjgustlr3@google.co.kr
# '''

# result = re.findall(r'[\w\.-]+@[\w\.-]+', test_string)
# print(result)

# 정규표현식 알아두기
# . : 하나의 문자와 일치한다
# [] : 순서와 상관없이 포함된 문자와 일치한다
# [^]: 포함되지 않은 문자 중 하나를 의미한다
# ^: 문자열의 시작 위치를 나타낸다. 여러 줄에서 처리하는 때에는 각 줄의 시작을 나타낸다.
# $: 문자열의 마지막 위치 또는 개행문자의 바로 앞 위치를 나타낸다
# (): 괄호 안의 일치되는 부븐을 묶어서 사용할 수 있다
# \1: 1~9까지의 숫자를 표현
# \w: 영어소문자, 언더바, 영어대문자, 숫자를 표현
# *: 바로 앞의 패턴이 0번 이상 일치 합니다.
# [1,2]: 바로 앞의 패턴이 최소 1번 최대 2번 일치합니다.
# ? : 바로 앞의 패턴이 0번 또는 1번 일치합니다.
# + : 바로 앞의 패턴이 1번 이상 일치합니다.
# | : 앞의 패턴 또는 뒤의 패턴 중 하나와 일치합니다.
# --------------------------------------------------------------------------------------------------------------------------------------

# 리스트에서 중복 내용 제거하는 로직

# import re

# test_string = '''
# tjgustlr@tjgustlr.com
# tjgustlr@naver.com
# tjgustlr1@yahoo.com
# tjgustlr2@daum.com
# tjgustlr3@google.co.kr
# '''

# result = re.findall(r'[\w\.-]+@[\w\.-]+', test_string)

# result = list(set(result)) # set을 사용하여 중복 제거
# print(result)
# --------------------------------------------------------------------------------------------------------------------------------------

# 사이트에서 이메일 수집하는 코드

import requests
import re

from sqlalchemy import true

url = 'https://news.v.daum.net/v/20220711092022897'

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

result = list(set(results))

print(results)
# --------------------------------------------------------------------------------------------------------------------------------------

# 수집한 이메일 주소를 엑셀에 저장하는 코드 만들기

import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://news.v.daum.net/v/20220711092022897'

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

result = list(set(results))

print(results)

try :
    wb = load_workbook(r'\.email.xlsx', data_only=true) # 엑셀 파일이 이미 존재한다면 읽어온다.
except: # 없다면 생성한다
    wb = Workbook()
    sheet = wb.active

for result in results: # 이메일을 수집한 수 만큼 반복해서 저장
    sheet.append([result])
    
wb.save(r'.\email.xlsx')