# 구글 및 네이버를 통해 파이썬 코드로 이메일을 보내고 엑셀 파일에 기록된 이메일 주소를 읽어 자동으로 대량의 이메일을 보내는 코드를 만든다
# 네이버로 보내기
# import smtplib
# from email.mime.text import MIMEText # 이메일 사용을 위한 라이브러리 모음

# send_email = '' 
# send_pwd = ''

# recv_email = 'rlaejrqo465@gmail.com'

# smtp_name = 'smtp.naver.com' # 네이버의 smtp 주소와 포트번호
# smtp_port = '587'

# text = '''          # 메일 내용
#         서현식 
#         한동근
# '''

# msg = MIMEText(text) # 메시지 형식을 텍스트 형식으로 지정

# msg['Subject'] = '메일제목'
# msg['From'] = send_email
# msg['To'] = recv_email
# print(msg.as_string())

# s = smtplib.SMTP( smtp_name, smtp_port) # 메일을 보냄
# s.starttls()
# s.login( send_email, send_pwd)
# s.sendmail(send_email,recv_email,msg.as_string())
# s.quit()
# ------------------------------------------------------------------------------------------------------------------------------

# 구글로 보내기

# import smtplib
# from email.mime.text import MIMEText

# send_email = ''
# send_pwd = ''

# recv_email = 'rlaejrqo465@naver.com'

# smtp_name = 'smtp.gmail.com'
# smtp_port = 587

# text = '''
#     한동근
#     서현식
# '''

# msg = MIMEText(text)

# msg['Subject'] = '메일제목'
# msg['From'] = send_email
# msg['To'] = recv_email
# print(msg.as_string())

# s = smtplib.SMTP( smtp_name, smtp_port) # 메일을 보냄
# s.starttls()
# s.login( send_email, send_pwd)
# s.sendmail(send_email,recv_email,msg.as_string())
# s.quit()
# ----------------------------------------------------------------------------------------------------------------------------------------------------------------------

# 파일을 첨부하여 메일 보내기

# from fileinput import filename
# import smtplib
# from email.mime.multipart import MIMEMultipart
# from email.mime.text import MIMEText
# from email.mime.application import MIMEApplication

# send_email = 'rlaejrqo465@gmail.com'
# send_pwd = 'tjabnrlzzmjqoczk'

# recv_email = 'rlaejrqo465@naver.com'

# smtp_name = 'smtp.gmail.com'
# smtp_port = '587'

# msg = MIMEMultipart() # 메일을 복합형식으로 선언해 첨부파일을 보낼 수 있게 함

# msg['Subject'] = '첨부파일 테스트입니다'
# msg['From'] = send_email
# msg['To'] = recv_email

# text = '''
#     서현식
#     한동근
#     강기모
# '''
# contentPart = MIMEText(text)
# msg.attach(contentPart)

# etc_file_path = r'./강원도날씨.png' # 첨부파일 경로 지정
# with open(etc_file_path, 'rb') as f: # 첨부파일을 읽어 파일을 첨부
#     etc_part = MIMEApplication(f.read())
#     etc_part.add_header('Content-Dispositon','attachment', filename='강원도날씨.png') 
#     msg.attach(etc_part)
    
# s = smtplib.SMTP( smtp_name, smtp_port)
# s.starttls()
# s.login( send_email, send_pwd)
# s.sendmail(send_email,recv_email,msg.as_string())
# s.quit()    
# ---------------------------------------------------------------------------------------------------------------------

# HTML 형식으로 보내기

# import smtplib
# from email.mime.multipart import MIMEMultipart
# from email.mime.text import MIMEText

# send_email = "보내는메일주소"
# send_pwd = "비밀번호"

# recv_email = "받는메일주소"

# smtp_name = "smtp.naver.com" 
# smtp_port = 587

# msg = MIMEMultipart()

# msg['Subject'] ="html로 보내는 메일 입니다."
# msg['From'] = send_email 
# msg['To'] = recv_email 

# html_body = """
# <p>안녕하세요 html 형식으로 보내는 이메일 테스트 입니다.</p>
# <p><span style="color: #0000ff;">글자의 색상을 지정하거나</span></p>
# <h1>크기를 조정할수 있습니다.</h1>
# <p>표도 만들수 있습니다.</p>
# <table style="height: 83px;" width="241">
# <tbody>
# <tr>
# <td style="width: 73px;">1</td>
# <td style="width: 73px;">2</td>
# <td style="width: 73px;">3</td>
# </tr>
# <tr>
# <td style="width: 73px;">표를</td>
# <td style="width: 73px;">만들수&nbsp;</td>
# <td style="width: 73px;">있습니다.</td>
# </tr>
# <tr>
# <td style="width: 73px;">4</td>
# <td style="width: 73px;">5</td>
# <td style="width: 73px;">6</td>
# </tr>
# </tbody>
# </table>
# <p>&nbsp;</p>
# """

# msg.attach( MIMEText(html_body,'html') ) 

# s=smtplib.SMTP( smtp_name , smtp_port )
# s.starttls()
# s.login( send_email , send_pwd )
# s.sendmail( send_email, recv_email, msg.as_string() )
# s.quit()
# ---------------------------------------------------------------------------------------------------------------------

# 엑셀파일을 읽어 메일 보내기

from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb = load_workbook(r'./이메일주소.xlsx', data_only=True) # 이메일 주소 엑셀파일을 읽어온다
load_ws = load_wb.active

for i in range(1, load_ws.max_row + 1): # 1부터 엑셀의 마지막 행까지 읽어 반복
    recv_email_value = load_ws.cell(i,1).value
    print('성공:', recv_email_value)
    
    try: # 에러가 없으면 메일을 보낸다.
        send_email = 'rlaejrqo465@naver.com'
        send_pwd = 'ghkdrldns1@'
        
        recv_email = recv_email_value
        
        smtp_name = 'smtp.naver.com'
        smtp_port = 587
        
        msg = MIMEMultipart()

        msg['Subject'] ="엑셀에서 메일 주소를 읽어 보냅니다"
        msg['From'] = send_email 
        msg['To'] = recv_email 
        
        text = '''
            메일내용
        '''
        
        msg.attach(MIMEText(text))
        
        s = smtplib.SMTP(smtp_name,smtp_port)
        s.starttls()
        s.login(send_email, send_pwd)
        s.sendmail(send_email,recv_email,msg.as_string())
        s.quit()
        
    except:
        print('에러:',recv_email_value)
        
        